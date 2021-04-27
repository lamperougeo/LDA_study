# -*- coding = utf-8 -*-
# @Time : 2021/4/4 9:52
# Author : ZERO
# @File : count_important.py
# @Software : PyCharm


import jieba
from gensim import corpora
import gensim
import openpyxl
import matplotlib.pyplot as plt
import zhconv
from matplotlib import pyplot as plt
from time import *


#统计运行程序时间
starttime = time()

#使用openpyxl读取文件
wb = openpyxl.load_workbook("allreviews.xlsx")  # 修改1  语料库xlsx文件路径
ws = wb.active
lines = []

for i in range(1,int(ws.max_row)+1):
    lines.append(ws.cell(row = i,column = 1 ).value)  #根据文本所在单元列修改column，默认为1


#打印文件长度
print("共有评论：%s条"%(len(lines)))


#将评论写入content中，为了将换行符去除、同时将繁体字转为简体
content = []
for each in lines:#将评论数据写入新的 list
    line = each.strip('\n')
    line_simple = zhconv.convert(line, 'zh-cn')
    content.append(line_simple)


#定义分词函数
def chinese_word_cut(mytext):
    return ' '.join(jieba.lcut(mytext))


#切分后加入seg中
seg = []
for each in content:
    seg.append(chinese_word_cut(each))

#加载停用词 hit_stopwords.txt
with open('hit_stopwords.txt','r',errors='ignore',encoding ='utf-8' ) as stop_fi:    #修改2 停用词路径
    lines = stop_fi.readlines()

    stopwords = [x.strip() for x in lines]


#将去除停用词后的切分结果写入texts中
texts =[[word for word in document.lower().split() if word not in stopwords] for document in seg]


#加载语料库
dictionary = corpora.Dictionary(texts)
corpus = [dictionary.doc2bow(text) for text in texts]

#检查数据读取是否准确
print('Number of unique tokens: %d' % len(dictionary))#读取到的词
print('Number of documents: %d' % len(corpus))#读取到的文件


#输入需要的主题数，以及每个主题下的关键词数
def lda(topic,words):
    #加载LDA模型
    ladmodel = gensim.models.ldamodel.LdaModel(corpus, num_topics =topic, id2word = dictionary, passes = 30 )

    #输出主题模型结果
    result = ladmodel.print_topics(num_topics=topic,num_words = words )
    print(result)

    #计算困惑度：
    perplexity = ladmodel.log_perplexity(corpus)

    cv_tmp = gensim.models.CoherenceModel(model=ladmodel, texts=texts, dictionary=dictionary, coherence='c_v')
    print(topic,perplexity)
    #返回主题数和困惑度
    return (topic,perplexity)
    # print(cv_tmp.get_coherence())

#困惑度绘图
def draw_picture(topic_min,topic_max,words):
    x = []
    y = []
    for i in range(topic_min,topic_max+1):
       bridge = lda(i,words)
       x.append(bridge[0])
       y.append(bridge[1])

    plt.plot(x,y)
    plt.show()

    
    

if __name__=="__main__":


    #计算LDA（填写主题数，每个主题下的主题词数）
    lda(5,10)


    #计算困惑度，并绘图 （填写最小主题数 最大主题数  每个主题下的主题词数） 不需要则直接注释掉
    draw_picture(5,10,3)

    endtime = time()
    print('该程序共耗时：%s s'%(endtime - starttime))


